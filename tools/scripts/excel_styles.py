from openpyxl.styles import PatternFill, Alignment, Border, Side, NamedStyle, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.colors import Color
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation

styles = []

align_center = Alignment(
    horizontal="center",
    vertical="center",
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=True,
    indent=0,
    wrapText=True,
)
align_left = Alignment(
    horizontal="general",
    vertical="center",
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=True,
    indent=0,
    justifyLastLine=True,
    wrapText=True,
)
FONT = "Avenir"

styles_metadata = [
    {"name": "text", "font": {'name': FONT, 'underline': "single", 'color': "1CA4FC"}, "alignment": "left", "background": ""},
    {"name": "center", "font": {'name': FONT}, "alignment": "center", "background": ""},
    {"name": "blue_link", "font": {'name': FONT, 'underline': "single", 'color': "1CA4FC"}, "alignment": "center", "background": ""},
    {"name": "gray", "font": {'name': FONT}, "alignment": "center", "background": "00C0C0C0"},
    {"name": "blue", "font": {'name': FONT}, "alignment": "center", "background": "0033CCCC"},
    {"name": "green", "font": {'name': FONT}, "alignment": "center", "background": "0099CC00"},
    {"name": "orange", "font": {'name': FONT}, "alignment": "center", "background": "00FF9900"},
]


def create_style(params):

    style = NamedStyle(name=params.get("name"))
    if params.get("font"):
        font_color = params.get("font").get('color')
        if font_color:
            params.get("font")['color'] = Color(rgb=params.get("font").get('color'))
        style.font = Font(**params.get("font"))

    # bd = Side(style='thick', color="FFFFFF")
    # style.border = Border(left=bd, top=bd, right=bd, bottom=bd)
    alignment = params.get("alignment")
    if alignment == "center":
        style.alignment = align_center
    else:
        style.alignment = align_left

    if params.get("background"):
        style.fill = PatternFill("solid", fgColor=params.get("background"))
        bd = Side(style="thick", color="FFFFFF")
        style.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    return style


def load_styles(wb):
    styles = [create_style(style) for style in styles_metadata]

    underline = NamedStyle(name="underline")
    underline.font = Font(name=FONT, size=15, bold=True, color="1CA4FC")
    bd = Side(style="medium", color="1CA4FC")
    underline.border = Border(bottom=bd)
    styles.append(underline)

    big_title = NamedStyle(name="big_title")
    big_title.font = Font(name=FONT, size=25)
    big_title.alignment = align_left
    styles.append(big_title)

    gray_header = NamedStyle(name="gray_header")
    gray_header.font = Font(name=FONT, bold=True, color="00C0C0C0")
    gray_header.alignment = align_center
    styles.append(gray_header)

    [wb.add_named_style(style) for style in styles]


# Conditional Formatting for STATUS
status_validation = DataValidation(type="list", formula1='"Pass,Fail,N/A"', allow_blank=True)

red_text = Font(color="9C0006")
red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(font=red_text, fill=red_fill, alignment=align_center)
rule_fail = Rule(type="containsText", operator="containsText", text="Fail", dxf=dxf)
rule_fail.formula = ['NOT(ISERROR(SEARCH("Fail",J11)))']

green_text = Font(color="38761D")
green_fill = PatternFill(bgColor="B6D7A8")
dxf = DifferentialStyle(font=green_text, fill=green_fill, alignment=align_center)
rule_pass = Rule(type="containsText", operator="containsText", text="Pass", dxf=dxf)
rule_pass.formula = ['NOT(ISERROR(SEARCH("Pass",J11)))']

gray_text = Font(color="666666")
gray_fill = PatternFill(bgColor="CCCCCC")
dxf = DifferentialStyle(font=gray_text, fill=gray_fill, alignment=align_center)
rule_na = Rule(type="containsText", operator="containsText", text="N/A", dxf=dxf)
rule_na.formula = ['NOT(ISERROR(SEARCH("N/A",J11)))']
