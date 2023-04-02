import yaml
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from enum import IntEnum

import combine_data_for_checklist


import excel_styles_and_validation

""" Tool for exporting the MASVS requirements as a checklist including MASTG coverage.

    By Carlos Holguera

    Copyright (c) 2022 OWASP Foundation

    Permission is hereby granted, free of charge, to any person obtaining a copy
    of this software and associated documentation files (the "Software"), to deal
    in the Software without restriction, including without limitation the rights
    to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
    copies of the Software, and to permit persons to whom the Software is
    furnished to do so, subject to the following conditions:

    The above copyright notice and this permission notice shall be included in all
    copies or substantial portions of the Software.

    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
    AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
    LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
    OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
    SOFTWARE.

"""

# TODO parametrize & create a function
# TODO read sheet, col ids and cell styles (centered, left, colored, character if true, etc) from yaml

MASVS_TITLES = {
    "MASVS-STORAGE": "Storage",
    "MASVS-CRYPTO": "Cryptography",
    "MASVS-AUTH": "Authentication",
    "MASVS-NETWORK": "Network Communication",
    "MASVS-PLATFORM": "Platform Interaction",
    "MASVS-CODE": "Code Quality",
    "MASVS-RESILIENCE": "Resilience",
}

MASVS = None
MASTGVERSION = ""
MASTGCOMMIT = ""
MASVSVERSION = ""
MASVSCOMMIT = ""

MAS_WEBSITE_ROOT = "https://owasp.org/mas"

class Position(IntEnum):
    ID = 2
    PLATFORM = 3
    TEXT = 4
    L1 = 5
    L2 = 6
    R = 7
    STATUS = 8

WS_BASE_CONFIG = {
    "start_row": 6,
    "columns": [
        {"col": "B", "position": Position.ID, "name": "MASVS-ID", "width": 25, "style": "gray_header"},
        {"col": "C", "position": Position.PLATFORM, "name": "Platform", "width": 15, "style": "gray_header"},
        {"col": "D", "position": Position.TEXT, "name": "Description",  "width": 80, "style": "gray_header"},
        {"col": "E", "position": Position.L1, "name": "L1", "width": 5, "style": "gray_header"},
        {"col": "F", "position": Position.L2, "name": "L2", "width": 5, "style": "gray_header"},
        {"col": "G", "position": Position.R, "name": "R", "width": 5, "style": "gray_header"},
        {"col": "K", "position": Position.STATUS, "name": "Status", "width": 10, "style": "gray_header"},
    ]
}


def write_header(ws):

    ws.row_dimensions[2].height = 65
    ws.merge_cells(start_row=2, end_row=4, start_column=2, end_column=3)

    img = Image("Document/Images/logo_circle.png")
    img.height = 140
    img.width = 140
    ws.add_image(img, "B2")

    img = Image("Document/Images/OWASP_logo-bw.png")
    img.height = img.height * 0.1
    img.width = img.width * 0.1
    ws.add_image(img, "F2")

    ws["D2"].value = "Mobile Application Security Checklist"
    ws["D2"].style = "big_title"

    ws["D3"].value = f'=HYPERLINK("https://github.com/OWASP/owasp-mastg/releases/tag/{MASTGVERSION}", "OWASP MASTG {MASTGVERSION} (commit: {MASTGCOMMIT})")'
    ws["D3"].font = Font(name=excel_styles_and_validation.FONT, color="00C0C0C0")
    ws["D4"].value = f'=HYPERLINK("https://github.com/OWASP/owasp-masvs/releases/tag/{MASVSVERSION}", "OWASP MASVS {MASVSVERSION} (commit: {MASVSCOMMIT})")'
    ws["D4"].font = Font(name=excel_styles_and_validation.FONT, color="00C0C0C0")

def set_columns_width(ws):
    for col in WS_BASE_CONFIG.get("columns"):
        ws.column_dimensions[col.get("col")].width = col.get("width")


def set_table_headers(row, ws):
    for col in WS_BASE_CONFIG["columns"]:
        ws.cell(row=row, column=col.get("position")).value = col.get("name")
        ws.cell(row=row, column=col.get("position")).style = col.get("style")


def write_title(ws, row, start_column, end_column, title):
    cell = ws.cell(row=row, column=start_column)
    cell.value = title
    cell.style = "underline"
    cell.alignment = excel_styles_and_validation.align_left

    ws.merge_cells(start_row=row, end_row=row, start_column=start_column, end_column=end_column)

    ws.row_dimensions[row].height = 25  # points

def create_security_requirements_sheet(wb):
    ws = wb.active
    ws.title = "Security Controls"
    ws.sheet_view.showGridLines = False
    write_header(ws)
    set_columns_width(ws)

    status_cells = 'H8:L400'
    ws.conditional_formatting.add(status_cells, excel_styles_and_validation.rule_fail)
    ws.conditional_formatting.add(status_cells, excel_styles_and_validation.rule_pass)
    ws.conditional_formatting.add(status_cells, excel_styles_and_validation.rule_na)

    row = WS_BASE_CONFIG["start_row"]

    for group_id, elements in MASVS.items():

        row = row + 1

        category_title = MASVS_TITLES[group_id]
        write_title(ws, row, Position.ID, Position.STATUS, category_title)
        row = row + 2

        set_table_headers(row, ws)
        row = row + 1
        
        ws.add_data_validation(excel_styles_and_validation.status_validation)
        row = row + 1

        # End header

        for element in elements:
            if element.get("MASVS-ID") != "":
                
                row = row + 1

                ws.cell(row=row, column=Position.ID).value = f'=HYPERLINK("{MAS_WEBSITE_ROOT}{element["path"]}", "{element["MASVS-ID"]}")'
                ws.cell(row=row, column=Position.ID).style = "center"

                ws.cell(row=row, column=Position.TEXT).value = element["Control / MASTG Test"]
                ws.cell(row=row, column=Position.TEXT).style = "text_bold"

                row = row + 2

            else:

                ws.cell(row=row, column=Position.PLATFORM).value = element["Platform"]
                ws.cell(row=row, column=Position.PLATFORM).style = "gray_text"

                ws.cell(row=row, column=Position.TEXT).value = f'=HYPERLINK("{MAS_WEBSITE_ROOT}{element["path"]}", "{element["Control / MASTG Test"]}")'
                ws.cell(row=row, column=Position.TEXT).style = "text"

                if element["L1"]:
                    ws.cell(row=row, column=Position.L1).style = "blue"
                if element["L2"]:
                    ws.cell(row=row, column=Position.L2).style = "green"
                if element["R"]:
                    ws.cell(row=row, column=Position.R).style = "orange"

                ws.row_dimensions[row].height = 55  # points

                status_cell = ws.cell(row=row, column=Position.STATUS).coordinate
                excel_styles_and_validation.status_validation.add(status_cell)

                row = row + 1
        
        row = row + 1

def create_about_sheet(wb):
    ws = wb.create_sheet("About")
    ws.sheet_view.showGridLines = False
    write_header(ws)
    set_columns_width(ws)

    row = WS_BASE_CONFIG["start_row"]
    first_col = WS_BASE_CONFIG["columns"][0].get("position")
    last_col = WS_BASE_CONFIG["columns"][-1].get("position")

    row = row + 2

    write_title(ws, row, first_col, last_col, "About the Project")

    row = row + 2

    ws.cell(row=row, column=first_col).value = "The OWASP Mobile Application Security (MAS) flagship project led by Carlos Holguera and Sven Schleier defines the industry standard for mobile application security."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = MAS_WEBSITE_ROOT
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    row = row + 2

    ws.cell(row=row, column=first_col).value = "The OWASP MASVS (Mobile Application Security Verification Standard) is a standard that establishes the security requirements for mobile app security."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://mas.owasp.org/MASTG/"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    row = row + 2

    ws.cell(row=row, column=first_col).value = "The OWASP MASTG (Mobile Application Security Testing Guide) is a comprehensive manual for mobile app security testing and reverse engineering. It describes technical processes for verifying the controls listed in the MASVS."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://mas.owasp.org/MASVS/"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    row = row + 2

    write_title(ws, row, first_col, last_col, "Feedback")

    row = row + 2

    ws.cell(row=row, column=first_col).value = "If you have any comments or suggestions, please post them on our GitHub Discussions."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://github.com/OWASP/owasp-mastg/discussions/categories/ideas"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    row = row + 2

    write_title(ws, row, first_col, last_col, "Licence")

    row = row + 2

    ws.cell(row=row, column=first_col).value = "Copyright © 2023 The OWASP Foundation. This work is licensed under a Creative Commons Attribution-ShareAlike 4.0 International License. For any reuse or distribution, you must make clear to others the license terms of this work."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://github.com/OWASP/owasp-mastg/blob/master/License.md"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'


def generate_spreadsheet(output_file):

    wb = Workbook()
    excel_styles_and_validation.load_styles(wb)

    create_security_requirements_sheet(wb)
    create_about_sheet(wb)

    wb.save(filename=output_file)


def main():
    global MASVS, LANG, MASTGVERSION, MASTGCOMMIT, MASVSVERSION, MASVSCOMMIT
    import argparse

    parser = argparse.ArgumentParser(description="Export the MAS checklist as Excel.")
    parser.add_argument("-o", "--outputfile", required=False, default="OWASP_MAS_Checklist.xlsx")
    parser.add_argument("-v1", "--mastgversion", required=False, default="x.x.x")
    parser.add_argument("-c1", "--mastgcommit", required=False, default="xxxxxx")
    parser.add_argument("-v2", "--masvsversion", required=False, default="y.y.y")
    parser.add_argument("-c2", "--masvscommit", required=False, default="yyyyyy")

    args = parser.parse_args()

    # set global vars
    MASVS = combine_data_for_checklist.get_checklist_dict()
    MASTGVERSION = args.mastgversion
    MASTGCOMMIT = args.mastgcommit
    MASVSVERSION = args.masvsversion
    MASVSCOMMIT = args.masvscommit

    print(f"Generating Checklist for MASTG {MASTGVERSION} ({MASTGCOMMIT}) and MASVS {MASVSVERSION} ({MASVSCOMMIT})")

    generate_spreadsheet(args.outputfile)


if __name__ == "__main__":
    main()
