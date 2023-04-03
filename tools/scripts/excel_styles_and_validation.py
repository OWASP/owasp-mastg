from openpyxl.styles import PatternFill, Alignment, Border, Side, NamedStyle, Font
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles.colors import Color
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation

styles = []

align_center = Alignment(
    horizontal="center",
    vertical="center",
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=True,
    indent=0,
    wrapText=True,
)
align_left = Alignment(
    horizontal="general",
    vertical="center",
    text_rotation=0,
    wrap_text=True,
    shrink_to_fit=True,
    indent=0,
    justifyLastLine=True,
    wrapText=True,
)
FONT = "Avenir"
MAS_BLUE = "499FFF"

MASVS_COLORS = {
    "MASVS-STORAGE": "DF5C8D",
    "MASVS-CRYPTO": "FABF50",
    "MASVS-AUTH": "317CC0",
    "MASVS-NETWORK": "EA5F33",
    "MASVS-PLATFORM": "F1A050",
    "MASVS-CODE": "5FACD3",
    "MASVS-RESILIENCE": "4FB991",
    "MASVS-PRIVACY": "8B5F9E",
}

styles_metadata = [
    {"name": "text", "font": {'name': FONT}, "alignment": "left", "background": ""},
    {"name": "text_bold", "font": {'name': FONT, 'bold': True}, "alignment": "left", "background": ""},
    {"name": "text_bold_medium", "font": {'name': FONT, 'size': 14, 'bold': True}, "alignment": "left", "background": ""},
    {"name": "center", "font": {'name': FONT}, "alignment": "center", "background": ""},
    {"name": "blue_link", "font": {'name': FONT, 'underline': "single", 'color': MAS_BLUE}, "alignment": "center", "background": ""},
    {"name": "gray", "font": {'name': FONT}, "alignment": "center", "background": "00C0C0C0"},
    {"name": "blue", "font": {'name': FONT}, "alignment": "center", "background": "0033CCCC"},
    {"name": "green", "font": {'name': FONT}, "alignment": "center", "background": "0099CC00"},
    {"name": "orange", "font": {'name': FONT}, "alignment": "center", "background": "00FF9900"},
    {"name": "underline", "font": {'name': FONT, 'size': 15, 'bold': True, 'color': MAS_BLUE}, "border": {"bottom": {"style": "medium", "color": MAS_BLUE}}, "alignment": align_left},
    {"name": "big_title", "font": {'name': FONT, 'size': 30, 'color': 'ffffff'}, "alignment": align_left},
    {"name": "medium_title", "font": {'name': FONT, 'size': 22, 'color': 'ffffff'}, "alignment": align_left},
    {"name": "gray_header", "font": {'name': FONT, 'size': 15, 'bold': True, 'color': "00C0C0C0"}, "alignment": align_center},
    {"name": "gray_text", "font": {'name': FONT, 'color': "00C0C0C0"}, "alignment": align_center},
    {"name": "versions_white", "font": {'name': FONT, 'size': 10, 'color': 'ffffff'}, "alignment": align_left},

]

for group in MASVS_COLORS:
    color_metadata = {
        "name": group, 
        "font": {'name': FONT, 'size': 15, 'bold': True, 'color': MASVS_COLORS[group]},
        "border": {'bottom': {'style': 'medium', 'color': MASVS_COLORS[group]}},
        "alignment": "left", 
        "background": MASVS_COLORS[group]}
    styles_metadata.append(color_metadata)

def create_style(params):

    style = NamedStyle(name=params.get("name"))
    if params.get("font"):
        style.font = Font(**params.get("font"))
    if params.get("border"):
        sides_dict = {}
        for key in params.get("border"):
            sides_dict[key] = Side(**params.get("border")[key])

        style.border = Border(**sides_dict)

    alignment = params.get("alignment")
    if alignment == "center":
        style.alignment = align_center
    else:
        style.alignment = align_left

    if params.get("background"):
        style.fill = PatternFill("solid", fgColor=params.get("background"))
        # bd = Side(style="thick", color="FFFFFF")
        # style.border = Border(left=bd, top=bd, right=bd, bottom=bd)

    return style


def load_styles(wb):
    styles = [create_style(style) for style in styles_metadata]

    [wb.add_named_style(style) for style in styles]


# Data Validation for STATUS

status_validation = DataValidation(type="list", formula1='"Pass,Fail,N/A"', allow_blank=True)

# Conditional Formatting for STATUS

red_text = Font(color="9C0006")
red_fill = PatternFill(bgColor="FFC7CE")
dxf = DifferentialStyle(font=red_text, fill=red_fill, alignment=align_center)
rule_fail = Rule(type="containsText", operator="containsText", text="Fail", dxf=dxf)
rule_fail.formula = ['NOT(ISERROR(SEARCH("Fail",I13)))']

green_text = Font(color="38761D")
green_fill = PatternFill(bgColor="B6D7A8")
dxf = DifferentialStyle(font=green_text, fill=green_fill, alignment=align_center)
rule_pass = Rule(type="containsText", operator="containsText", text="Pass", dxf=dxf)
rule_pass.formula = ['NOT(ISERROR(SEARCH("Pass",I13)))']

gray_text = Font(color="666666")
gray_fill = PatternFill(bgColor="CCCCCC")
dxf = DifferentialStyle(font=gray_text, fill=gray_fill, alignment=align_center)
rule_na = Rule(type="containsText", operator="containsText", text="N/A", dxf=dxf)
rule_na.formula = ['NOT(ISERROR(SEARCH("N/A",I13)))']
