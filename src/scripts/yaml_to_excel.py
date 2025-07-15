import yaml
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill
from enum import IntEnum

import combine_data_for_checklist


import excel_styles_and_validation as mas_styles

""" Tool for exporting the MASVS requirements as a checklist including MASTG coverage.

    By Carlos Holguera

    Copyright (c) 2023 OWASP Foundation

    Permission is hereby granted, free of charge, to any person obtaining a copy
    of this software and associated documentation files (the "Software"), to deal
    in the Software without restriction, including without limitation the rights
    to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
    copies of the Software, and to permit persons to whom the Software is
    furnished to do so, subject to the following conditions:

    The above copyright notice and this permission notice shall be included in all
    copies or substantial portions of the Software.

    THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
    IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
    FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
    AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
    LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
    OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
    SOFTWARE.

"""

# TODO parametrize & create a function
# TODO read sheet, col ids and cell styles (centered, left, colored, character if true, etc) from yaml

MASVS_GROUPS = None

MASVS = None
MASTGVERSION = ""
MASTGCOMMIT = ""
MASVSVERSION = ""
MASVSCOMMIT = ""

MAS_WEBSITE_ROOT = "https://mas.owasp.org/"

class Position(IntEnum):
    START = 1
    ID = 2
    PLATFORM = 3
    TEXT = 4
    L1 = 5
    L2 = 6
    R = 7
    PADDING = 8
    STATUS = 9
    EMPTY = 10

WS_BASE_CONFIG = {
    "start_row": 7,
    "columns": [
        {"col": "A", "position": Position.START, "name": "", "width": 10, "style": "gray_header"},
        {"col": "B", "position": Position.ID, "name": "MASVS-ID", "width": 25, "style": "gray_header"},
        {"col": "C", "position": Position.PLATFORM, "name": "Platform", "width": 15, "style": "gray_header"},
        {"col": "D", "position": Position.TEXT, "name": "Description",  "width": 80, "style": "gray_header"},
        {"col": "E", "position": Position.L1, "name": "L1", "width": 5, "style": "gray_header"},
        {"col": "F", "position": Position.L2, "name": "L2", "width": 5, "style": "gray_header"},
        {"col": "G", "position": Position.R, "name": "R", "width": 5, "style": "gray_header"},
        {"col": "H", "position": Position.PADDING, "name": "", "width": 5, "style": "gray_header"},
        {"col": "I", "position": Position.STATUS, "name": "Status", "width": 10, "style": "gray_header"},
        {"col": "J", "position": Position.EMPTY, "name": "", "width": 15, "style": "gray_header"},
    ]
}


def write_header(ws, category_title, background_color="ffffff", font_color="000000"):

    ws.row_dimensions[2].height = 65


    img = Image("Document/Images/logo_circle.png")
    img.height = 165
    img.width = 165
    ws.add_image(img, "B2")

    img = Image("Document/Images/OWASP_logo_white.png")
    img.height = img.height * 0.15
    img.width = img.width * 0.15
    ws.add_image(img, "F2")

    ws["C2"].value = "Mobile Application Security Checklist"
    ws["C2"].style = "big_title"

    ws["C3"].value = category_title
    ws["C3"].style = "medium_title"

    ws["C5"].value = f'OWASP MASTG {MASTGVERSION} (commit: {MASTGCOMMIT})' + "    " + f'OWASP MASVS {MASVSVERSION} (commit: {MASVSCOMMIT})'
    ws["C5"].font = Font(name=mas_styles.FONT, color="ffffff")
    ws["C5"].style = "versions_white"

    # MAS logo
    ws.merge_cells(start_row=2, end_row=4, start_column=2, end_column=2)
    # OWASP logo
    ws.merge_cells(start_row=2, end_row=2, start_column=5, end_column=7)
    # Checklist Title
    ws.merge_cells(start_row=2, end_row=2, start_column=3, end_column=4)
    # Sheet title
    ws.merge_cells(start_row=3, end_row=3, start_column=3, end_column=4)
    # Version Info
    ws.merge_cells(start_row=5, end_row=5, start_column=3, end_column=4)

    # Set background color
    for col in WS_BASE_CONFIG.get("columns"):
        for row in range(1, 7):
            ws.cell(row=row, column=col.get("position")).fill = PatternFill(start_color=background_color, fill_type="solid")
 
def set_columns_width(ws):
    for col in WS_BASE_CONFIG.get("columns"):
        ws.column_dimensions[col.get("col")].width = col.get("width")


def set_table_headers(row, ws):
    for col in WS_BASE_CONFIG["columns"]:
        ws.cell(row=row, column=col.get("position")).value = col.get("name")
        ws.cell(row=row, column=col.get("position")).style = col.get("style")


def write_title(ws, row, start_column, end_column, title, masvs_font_style="underline"):
    cell = ws.cell(row=row, column=start_column)
    cell.value = title
    cell.style = masvs_font_style
    cell.alignment = mas_styles.align_left

    ws.merge_cells(start_row=row, end_row=row, start_column=start_column, end_column=end_column)

    ws.row_dimensions[row].height = 25  # points

def create_security_requirements_sheet(wb):
    first_sheet = wb.active

    for group_id, elements in MASVS.items():
        ws = wb.create_sheet(group_id)

        ws.title = group_id
        ws.sheet_view.showGridLines = False
        category_title = f"{group_id}: {MASVS_GROUPS[group_id]['title']}"

        write_header(ws, category_title, mas_styles.MASVS_COLORS[group_id])
        set_columns_width(ws)

        status_cells = 'I11:I400'
        ws.conditional_formatting.add(status_cells, mas_styles.rule_fail)
        ws.conditional_formatting.add(status_cells, mas_styles.rule_pass)
        ws.conditional_formatting.add(status_cells, mas_styles.rule_na)

        row = WS_BASE_CONFIG["start_row"]


        row = row + 1

        # category_title = f"{group_id}: {MASVS_GROUPS[group_id]['title']}"
        # write_title(ws, row, Position.ID, Position.STATUS, category_title, masvs_font_style=group_id)
        # row = row + 2

        set_table_headers(row, ws)
        row = row + 1
        
        ws.add_data_validation(mas_styles.status_validation)
        row = row + 1

        # End header

        for element in elements:
            # MASVS control
            if element.get("MASVS-ID") != "":
                
                row = row + 1

                ws.cell(row=row, column=Position.ID).value = f'=HYPERLINK("{MAS_WEBSITE_ROOT}{element["path"]}", "{element["MASVS-ID"]}")'
                ws.cell(row=row, column=Position.ID).style = "text_bold"

                ws.cell(row=row, column=Position.TEXT).value = element["Control / MASTG Test"]
                ws.cell(row=row, column=Position.TEXT).style = "text_bold"

                status_cell = ws.cell(row=row, column=Position.STATUS).coordinate
                mas_styles.status_validation.add(status_cell)

                row = row + 2
            # MASTG test
            elif element.get("Platform") != "":

                ws.cell(row=row, column=Position.PLATFORM).value = element["Platform"]
                ws.cell(row=row, column=Position.PLATFORM).style = "gray_text"

                ws.cell(row=row, column=Position.TEXT).value = f'=HYPERLINK("{MAS_WEBSITE_ROOT}{element["path"]}", "{element["Control / MASTG Test"]}")'
                ws.cell(row=row, column=Position.TEXT).style = "text"

                if element["L1"]:
                    ws.cell(row=row, column=Position.L1).style = "blue"
                if element["L2"]:
                    ws.cell(row=row, column=Position.L2).style = "green"
                if element["R"]:
                    ws.cell(row=row, column=Position.R).style = "orange"

                ws.row_dimensions[row].height = 55  # points

                status_cell = ws.cell(row=row, column=Position.STATUS).coordinate
                mas_styles.status_validation.add(status_cell)

                row = row + 1
        
        row = row + 1
        ws.cell(row=row, column=1).value = ""

    del wb[first_sheet.title]

def create_about_sheet(wb):
    ws = wb.create_sheet("About")
    ws.sheet_view.showGridLines = False
    write_header(ws, "About", background_color=mas_styles.MAS_BLUE)
    set_columns_width(ws)

    row = WS_BASE_CONFIG["start_row"]
    first_col = WS_BASE_CONFIG["columns"][1].get("position")
    last_col = WS_BASE_CONFIG["columns"][-2].get("position")

    row = row + 2

    write_title(ws, row, first_col, last_col, "About the Project")

    row = row + 2

    ws.cell(row=row, column=first_col).value = "The OWASP Mobile Application Security (MAS) flagship project led by Carlos Holguera and Sven Schleier \ndefines the industry standard for mobile application security."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = MAS_WEBSITE_ROOT
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    row = row + 2

    ws.cell(row=row, column=first_col).value = "The OWASP MASVS (Mobile Application Security Verification Standard) is a standard that establishes the \nsecurity requirements for mobile app security."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://mas.owasp.org/MASVS/"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    ws.cell(row=row, column=first_col+2).value = f'=HYPERLINK("https://github.com/OWASP/masvs/releases/tag/{MASVSVERSION}", "OWASP MASVS {MASVSVERSION} (commit: {MASVSCOMMIT})")'
    ws.cell(row=row, column=first_col+2).style = "text"

    row = row + 2

    ws.cell(row=row, column=first_col).value = "The OWASP MASTG (Mobile Application Security Testing Guide) is a comprehensive manual for mobile app security testing \nand reverse engineering. It describes technical processes for verifying the controls listed in the MASVS."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://mas.owasp.org/MASTG/"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    ws.cell(row=row, column=first_col+2).value = f'=HYPERLINK("https://github.com/OWASP/mastg/releases/tag/{MASTGVERSION}", "OWASP MASTG {MASTGVERSION} (commit: {MASTGCOMMIT})")'
    ws.cell(row=row, column=first_col+2).style = "text"


    row = row + 2

    write_title(ws, row, first_col, last_col, "Feedback")

    row = row + 2

    ws.cell(row=row, column=first_col).value = "If you have any comments or suggestions, please post them on our GitHub Discussions."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://github.com/OWASP/mastg/discussions/categories/ideas"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    row = row + 2

    write_title(ws, row, first_col, last_col, "Licence")

    row = row + 2

    ws.cell(row=row, column=first_col).value = "Copyright © 2023 The OWASP Foundation. This work is licensed under a Creative Commons Attribution-ShareAlike 4.0 International License. \nFor any reuse or distribution, you must make clear to others the license terms of this work."
    ws.merge_cells(start_row=row, end_row=row, start_column=first_col, end_column=last_col)
    ws.cell(row=row, column=first_col).style = "text"

    row = row + 2
    url = "https://github.com/OWASP/mastg/blob/master/License.md"
    ws.cell(row=row, column=first_col).value = f'=HYPERLINK("{url}", "{url}")'

    # padding
    row = row + 1
    ws.cell(row=row, column=first_col).value = ""
    row = row + 1
    ws.cell(row=row, column=first_col).value = ""

def generate_spreadsheet(output_file):

    wb = Workbook()
    mas_styles.load_styles(wb)

    create_security_requirements_sheet(wb)
    create_about_sheet(wb)

    wb.save(filename=output_file)


def main():
    global MASVS, MASVS_GROUPS, MASTGVERSION, MASTGCOMMIT, MASVSVERSION, MASVSCOMMIT
    import argparse

    parser = argparse.ArgumentParser(description="Export the MAS checklist as Excel.")
    parser.add_argument("-o", "--outputfile", required=False, default="OWASP_MAS_Checklist.xlsx")
    parser.add_argument("-v1", "--mastgversion", required=False, default="x.x.x")
    parser.add_argument("-c1", "--mastgcommit", required=False, default="xxxxxx")
    parser.add_argument("-v2", "--masvsversion", required=False, default="y.y.y")
    parser.add_argument("-c2", "--masvscommit", required=False, default="yyyyyy")

    args = parser.parse_args()

    # set global vars
    MASVS = combine_data_for_checklist.get_checklist_dict()
    MASVS_GROUPS = combine_data_for_checklist.get_masvs_groups()
    MASTGVERSION = args.mastgversion
    MASTGCOMMIT = args.mastgcommit
    MASVSVERSION = args.masvsversion
    MASVSCOMMIT = args.masvscommit

    print(f"Generating Checklist for MASTG {MASTGVERSION} ({MASTGCOMMIT}) and MASVS {MASVSVERSION} ({MASVSCOMMIT})")

    generate_spreadsheet(args.outputfile)

    print(f"Output file: {args.outputfile}")

if __name__ == "__main__":
    main()
